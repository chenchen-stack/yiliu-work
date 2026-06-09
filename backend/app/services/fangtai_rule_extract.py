"""从方太《收入/回款异常问题登记表》Excel 提取排查规则。"""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import BinaryIO

import openpyxl

DATA_DIR = Path(__file__).resolve().parents[2] / "data"
PRESET_JSON = DATA_DIR / "fangtai_troubleshooting_rules.json"


def _cell_str(v) -> str:
    if v is None:
        return ""
    if hasattr(v, "isoformat"):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def _infer_rule_type(text: str) -> str:
    if any(k in text for k in ("重复", "多行", "两套", "去重")):
        return "duplicate_record"
    if any(k in text for k in ("映射", "MDM", "抬头", "编码", "主数据", "发票类型")):
        return "mapping_anomaly"
    if any(k in text for k in ("回款", "付款", "分款", "收款", "回款单", "付款申请")):
        return "payment_mismatch"
    if any(k in text for k in ("接口", "回传", "同步", "传输", "下发报错", "未更新", "未回传", "接收成功")):
        return "sync_failure"
    if any(k in text for k in ("状态", "过账", "开票状态", "ZTSD", "不一致", "过账中", "开票中", "未开票")):
        return "status_mismatch"
    if any(k in text for k in ("帆软", "DRP", "LTC", "四列", "汇总")):
        return "fanruan_summary"
    return "amount_mismatch"


def _infer_severity(count: int | None, text: str) -> str:
    if count and count >= 50:
        return "high"
    if any(k in text for k in ("接口报错", "状态不一致", "未回传", "失败")):
        return "high"
    if count and count >= 10:
        return "medium"
    return "medium"


def parse_registration_sheet(ws) -> list[dict]:
    items: list[dict] = []
    current_category = ""
    current_group = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = [_cell_str(c) for c in row]
        if not any(cells):
            continue
        col0 = cells[0] if len(cells) > 0 else ""
        col1 = cells[1] if len(cells) > 1 else ""
        if col0 in ("收入", "回款", "合计"):
            current_category = col0
        if col1 and col1 not in ("问题归类", "系统配置问题", "系统流程问题", "行为操作问题", "问题提示归类"):
            if len(col1) < 40 and not col1.startswith("DMS") and not col1.startswith("SAP"):
                current_group = col1
        detail = cells[2] if len(cells) > 2 else ""
        if not detail or detail in ("问题明细描述",):
            continue
        if detail.startswith("系统") and len(detail) < 12:
            continue
        count_raw = cells[3] if len(cells) > 3 else ""
        count: int | None = int(count_raw) if count_raw.isdigit() else None
        cause = cells[4] if len(cells) > 4 else ""
        remedy = cells[5] if len(cells) > 5 else ""
        steps = cells[7] if len(cells) > 7 else ""
        if len(cells) > 19 and not steps:
            steps = cells[-1]
        blob = " ".join([detail, cause, remedy, steps])
        items.append({
            "source": "registration",
            "category": current_category or "收入",
            "problem_group": current_group,
            "name": detail[:60] + ("…" if len(detail) > 60 else ""),
            "problem_detail": detail,
            "count": count,
            "cause_category": cause,
            "remedy": remedy,
            "troubleshooting_steps": steps,
            "rule_type": _infer_rule_type(blob),
            "severity": _infer_severity(count, blob),
            "enabled": True,
        })
    return items


def parse_exception_sheet(ws, doc_type: str) -> list[dict]:
    items: list[dict] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        cells = [_cell_str(c) for c in row]
        if len(cells) < 8:
            continue
        seq = cells[1] if len(cells) > 1 else ""
        if not seq.isdigit():
            continue
        detail = cells[6] if len(cells) > 6 else ""
        remedy = cells[7] if len(cells) > 7 else ""
        if not detail:
            continue
        blob = f"{detail} {remedy}"
        items.append({
            "source": "exception_ticket",
            "category": doc_type,
            "problem_group": doc_type,
            "name": (detail[:56] + "…") if len(detail) > 56 else detail,
            "problem_detail": detail,
            "count": 1,
            "cause_category": cells[5] if len(cells) > 5 else "",
            "remedy": remedy,
            "troubleshooting_steps": remedy,
            "rule_type": _infer_rule_type(blob),
            "severity": "medium",
            "enabled": True,
        })
    return items


def extract_workbook_from_path(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return _extract_from_workbook(wb, path.name)
    finally:
        wb.close()


def extract_workbook_from_stream(stream: BinaryIO, filename: str) -> dict:
    wb = openpyxl.load_workbook(stream, read_only=True, data_only=True)
    try:
        return _extract_from_workbook(wb, filename)
    finally:
        wb.close()


def _extract_from_workbook(wb, source_name: str) -> dict:
    all_items: list[dict] = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        if "登记" in sn:
            all_items.extend(parse_registration_sheet(ws))
        elif "异常结算" in sn:
            all_items.extend(parse_exception_sheet(ws, "异常结算单"))
        elif "异常回款" in sn:
            all_items.extend(parse_exception_sheet(ws, "异常回款单"))
    seen: set[str] = set()
    unique: list[dict] = []
    for it in all_items:
        key = re.sub(r"\s+", "", it["problem_detail"])[:80]
        if key in seen:
            continue
        seen.add(key)
        unique.append(it)
    by_type: dict[str, list[dict]] = {}
    for it in unique:
        by_type.setdefault(it["rule_type"], []).append(it)
    type_labels = {
        "amount_mismatch": "金额差异",
        "duplicate_record": "重复数据",
        "mapping_anomaly": "主数据/映射异常",
        "status_mismatch": "状态不一致",
        "sync_failure": "接口/同步异常",
        "payment_mismatch": "回款差异",
        "fanruan_summary": "帆软汇总差异",
    }
    consolidated = []
    for rtype, group in by_type.items():
        steps_lines = []
        for i, g in enumerate(group[:8], 1):
            line = f"{i}. [{g.get('problem_group') or g.get('category')}] {g['problem_detail'][:120]}"
            if g.get("troubleshooting_steps"):
                line += f"\n   排查：{str(g['troubleshooting_steps'])[:200]}"
            steps_lines.append(line)
        consolidated.append({
            "rule_type": rtype,
            "name": f"方太·{type_labels.get(rtype, rtype)}排查规则",
            "condition": (
                f"基于方太《收入/回款异常问题登记表》沉淀：本类共 {len(group)} 条典型场景。"
                "核心判定见引擎逻辑；下列为财务实操排查要点。"
            ),
            "severity": "high" if rtype != "duplicate_record" else "medium",
            "threshold": 0 if rtype == "amount_mismatch" else None,
            "troubleshooting_steps": "\n".join(steps_lines),
            "sample_count": len(group),
            "samples": group[:5],
        })
    return {
        "title": "方太收入/回款异常排查规则（从登记表提取）",
        "source_file": source_name,
        "extracted_at": datetime.utcnow().isoformat(timespec="seconds"),
        "total_patterns": len(unique),
        "consolidated_rules": consolidated,
        "patterns": unique,
    }


def load_preset() -> dict:
    if not PRESET_JSON.exists():
        raise FileNotFoundError(f"预设规则文件不存在: {PRESET_JSON}")
    return json.loads(PRESET_JSON.read_text(encoding="utf-8"))
